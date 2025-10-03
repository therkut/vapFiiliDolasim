import os
import time
import pandas as pd
from selenium import webdriver
from openpyxl import load_workbook
from datetime import datetime, timedelta
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# === AYARLAR ===
DOWNLOAD_DIR = os.path.join(os.getcwd(), "data")
URL = "https://www.vap.org.tr/api/all-companies"

# === YARDIMCI FONKSİYONLAR ===

def get_target_date():
    """
    Veri çekme gününü belirler:
    - Pazartesi: Cuma gününün verisi alınır (3 gün önce)
    - Salı–Cuma: Bir önceki gün
    - Cumartesi–Pazar: Çalışmaz (None döner)
    """
    today = datetime.today()
    weekday = today.weekday()

    if weekday == 0:  # Pazartesi
        return today - timedelta(days=3)
    elif 1 <= weekday <= 4:  # Salı–Cuma
        return today - timedelta(days=1)
    else:  # Hafta sonu
        return None

def setup_driver(download_dir):
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "directory_upgrade": True
    }
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    return webdriver.Chrome(options=options)

def wait_for_download(download_dir, extensions=(".xls", ".xlsx"), timeout=30):
    """
    İndirilen dosyanın tamamlanmasını bekler.
    """
    for _ in range(timeout):
        files = [f for f in os.listdir(download_dir) if f.endswith(extensions)]
        if files:
            # En son indirilen dosyayı döndürür
            return max(files, key=lambda f: os.path.getctime(os.path.join(download_dir, f)))
        time.sleep(1)
    return None

def convert_xls_to_html(xls_path, html_path):
    """
    Excel dosyasını okuyup temizlenmiş HTML formatında kaydeder.
    Sayısal değerler yerel biçime çevrilir.
    """
    wb = load_workbook(xls_path, data_only=True)
    ws = wb.active

    def format_value(cell):
        val = cell.value
        if val is None:
            return ""
        if isinstance(val, (int, float)):
            fmt = cell.number_format
            try:
                if "0.00" in fmt or "0,00" in fmt:
                    # Türkçe format: 1.234,56 gibi
                    s = f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    return s
                return f"{int(val):,}".replace(",", ".")
            except:
                return str(val)
        return str(val)

    data = [
        [format_value(cell) for cell in row]
        for row in ws.iter_rows()
    ]

    # Boş satırları kaldır
    data = [row for row in data if any(cell != "" for cell in row)]

    # Boş sütunları kaldır
    if data:
        cols = len(data[0])
        non_empty_cols = [i for i in range(cols) if any(row[i] != "" for row in data)]
        data = [[row[i] for i in non_empty_cols] for row in data]

    df = pd.DataFrame(data[1:], columns=data[0])
    df.to_html(html_path, index=False, border=1, na_rep="")

# === ANA FONKSİYON ===

def main():
    target_date = get_target_date()
    if target_date is None:
        print("🛑 Hafta sonu. Script çalışmayacak.")
        return

    date_str = target_date.strftime("%d/%m/%Y")
    print(f"📅 Hedef veri tarihi: {date_str}")

    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    driver = setup_driver(DOWNLOAD_DIR)

    try:
        print("🌐 Sayfa açılıyor...")
        driver.get(URL)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input#datepicker"))
        )

        # Tarihi ayarla ve değişikliği tetikle
        driver.execute_script(f"""
            var input = document.querySelector("#datepicker");
            input.value = '{date_str}';
            input.dispatchEvent(new Event('change', {{ bubbles: true }}));
        """)

        driver.find_element(By.CSS_SELECTOR, "input.submit-btn").click()
        print("📥 İndirme işlemi başlatıldı...")

        downloaded_file = wait_for_download(DOWNLOAD_DIR)
        if not downloaded_file:
            print("❌ Dosya indirilemedi (süre doldu).")
            return

    finally:
        driver.quit()

    xls_path = os.path.join(DOWNLOAD_DIR, downloaded_file)

    # Tarihli ve sabit isimli HTML dosyalarının yolları
    date_suffix = target_date.strftime("-%d-%m-%Y")
    dated_html_path = os.path.join(DOWNLOAD_DIR, f"Fiili_Dolasim_Raporu_MKK{date_suffix}.html")
    fixed_html_path = os.path.join(DOWNLOAD_DIR, "Fiili_Dolasim_Raporu_MKK.html")

    print(f"📄 İndirilen dosya: {downloaded_file}")
    print("🔄 HTML'e dönüştürülüyor...")

    try:
        convert_xls_to_html(xls_path, dated_html_path)
        convert_xls_to_html(xls_path, fixed_html_path)

        print(f"✅ HTML dosyaları oluşturuldu:")
        print(f"  - {os.path.basename(dated_html_path)}")
        print(f"  - {os.path.basename(fixed_html_path)}")

    except Exception as e:
        print(f"❌ HTML dönüşümünde hata: {e}")

# === PROGRAM BAŞLANGICI ===

if __name__ == "__main__":
    main()
